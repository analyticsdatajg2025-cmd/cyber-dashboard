import requests
from datetime import date
from openpyxl import load_workbook

with open("mi_token.txt", "r", encoding="utf-8") as f:
    TOKEN = f.read().strip()

ARCHIVO_PROY = "SESIONES CONECTA - CYBER 2026 ABRIL.xlsx"
HEADERS = {"Authorization": f"OAuth {TOKEN}"}

MARCAS = {
    "LA CURACAO":  {"contador": "98373248",
                    "dias": {0:(2,3),1:(6,7),2:(10,11),3:(14,15),4:(18,19),5:(22,23),6:(24,25)}},
    "TIENDAS EFE": {"contador": "98373144",
                    "dias": {0:(28,29),1:(32,33),2:(36,37),3:(40,41),4:(44,45),5:(48,49),6:(50,51)}},
}
DIAS = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

def leer_proyecciones(col_proy):
    wb = load_workbook(ARCHIVO_PROY, data_only=True)
    ws = wb.active
    return [float(ws.cell(row=3+h, column=col_proy).value or 0) for h in range(24)]

def traer_sesiones_hora(contador):
    URL = "https://api-metrika.yandex.net/stat/v1/data/bytime"
    p = {"ids": contador, "metrics": "ym:s:visits",
         "date1": "today", "date2": "today", "group": "hour"}
    d = requests.get(URL, headers=HEADERS, params=p).json()
    serie = [int(v or 0) for v in d["data"][0]["metrics"][0]]
    corte = d.get("last_period_index", 23)
    return serie, corte

def traer_funnel(contador):
    URL = "https://api-metrika.yandex.net/stat/v1/data"
    metrics = ["ym:s:visits","ym:s:ecommerceRevenue","ym:s:ecommercePurchases",
               "ym:s:productImpressions","ym:s:productBasketsQuantity",
               "ym:s:productPurchasedQuantity"]
    p = {"ids": contador, "metrics": ",".join(metrics),
         "date1": "today", "date2": "today"}
    t = requests.get(URL, headers=HEADERS, params=p).json()["totals"]
    sesiones, ingresos, compras, vistos, carrito, comprados = t
    return {
        "ingresos": ingresos,
        "compras": int(compras),
        "vistos": int(vistos),
        "carrito": int(carrito),
        "comprados": int(comprados),
        "cr": (compras / sesiones) if sesiones else 0,        # compras / sesiones
        "ticket": (ingresos / compras) if compras else 0,     # ingresos / compras
    }

def armar_reporte(marca):
    cfg = MARCAS[marca]
    dia_idx = date.today().weekday()
    col_proy, _ = cfg["dias"][dia_idx]
    proy = leer_proyecciones(col_proy)
    real, corte = traer_sesiones_hora(cfg["contador"])
    funnel = traer_funnel(cfg["contador"])
    real_acum = sum(real[:corte+1])
    meta = sum(proy)
    return {
        "marca": marca, "dia": DIAS[dia_idx], "fecha": str(date.today()),
        "corte": corte, "proy": proy, "real": real,
        "meta": meta, "real_acum": real_acum,
        "avance": (real_acum / meta) if meta else 0,
        "funnel": funnel,
    }

def imprimir(r):
    print(f"\n{'='*50}\n  {r['marca']}  |  {r['dia']} {r['fecha']}  |  corte hora {r['corte']}:00\n{'='*50}")
    print("  --- SESIONES (corte) ---")
    print(f"  Meta del día : {r['meta']:>12,.0f}")
    print(f"  Real al corte: {r['real_acum']:>12,}")
    print(f"  % Avance     : {r['avance']:>12.1%}")
    f = r["funnel"]
    print("  --- FUNNEL (acumulado hoy) ---")
    print(f"  Ingresos Gross    : S/ {f['ingresos']:>12,.2f}")
    print(f"  CR Gross          : {f['cr']:>12.2%}")
    print(f"  Ticket Promedio   : S/ {f['ticket']:>10,.2f}")
    print(f"  Artículos vistos  : {f['vistos']:>12,}")
    print(f"  Añadidos carrito  : {f['carrito']:>12,}")
    print(f"  Artículos comprados:{f['comprados']:>12,}")
    print(f"  (compras/órdenes) : {f['compras']:>12,}")

if __name__ == "__main__":
    for marca in MARCAS:
        imprimir(armar_reporte(marca))