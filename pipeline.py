"""
pipeline.py — EL CEREBRO. Es lo unico que corre el cron.

DOS MODOS, DECIDIDOS SOLO POR LA FECHA
--------------------------------------
  CYBER   -> hoy cae dentro de CYBER_DAYS / CYBER_WOW.
             Corre cada hora. Proyeccion horaria del Excel. JUNTOZ visible.
             Escribe Sheets BACKEND + META HORARIA. Manda PNG + correo.
  DIARIO  -> cualquier otro dia del anio.
             Corre 1 vez al dia. Solo meta diaria (col F del Sheet mensual),
             sin curva horaria. JUNTOZ excluido. Sin correo. Sin Sheets.

Para activar el modo cyber SOLO se editan CYBER_DAYS / CYBER_WOW abajo.

La variable de entorno MODO dice QUE WORKFLOW esta corriendo, no que hacer:
  MODO=diario -> si hoy ES cyber, el job se salta (lo cubre el horario).
  MODO=cyber  -> si hoy NO es cyber, el job se salta.
  MODO=auto   -> corre siempre con el modo que toque (util para probar local).

Flujo de una corrida:
  1. Extrae de Yandex (por marca): funnel, sesiones/hora, ingresos/hora,
     campanias, top productos.
  2. Proyeccion: Excel horario (cyber) o Sheet mensual col F (diario).
  3. Escribe data.json + historico.json + neto.json (bloque por_dia).
  4. Solo en cyber: Sheets BACKEND / META HORARIA + PNG + correo.

Si algo falla: correo de alerta y sale con codigo != 0.
"""
import json
import sys
import traceback
import requests
import os
from proy_diaria import (leer_proyeccion_diaria, leer_neto_diario,
                         leer_meta_mensual)
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

RE_CAMPANA_NUMERICA = re.compile(r"^\d{15,22}$") 

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
with open("mi_token.txt", "r", encoding="utf-8") as f:
    TOKEN = f.read().strip()

HEADERS = {"Authorization": f"OAuth {TOKEN}"}
URL_DATA   = "https://api-metrika.yandex.net/stat/v1/data"
URL_BYTIME = "https://api-metrika.yandex.net/stat/v1/data/bytime"

# UN SOLO juego de archivos para los dos modos. El dashboard lee siempre estos.
SALIDA_JSON      = "data.json"
SALIDA_HISTORICO = "historico.json"
ARCHIVO_NETO     = "neto.json"

HIST_INICIO = date(2026, 7, 6)     # ancla del historico (no baja de aqui)
HIST_RETENCION_DIAS = 60           # dias cerrados que se conservan
HIST_REFRESH_DIAS = 3              # dias cerrados recientes que se vuelven a
                                   # consultar en cada corrida diaria: Yandex
                                   # sigue asentando datos por 1-2 dias, asi el
                                   # historico deja de quedar congelado con la
                                   # foto del primer dia.
BACKFILL_MAX = 7                   # dias faltantes que se rellenan por corrida
                                   # (protege la cuota de Yandex al volver de
                                   #  un hueco largo, ej. JUNTOZ al entrar cyber)
ARCHIVO_PROY = "PROYECCIONES_CYBER_JULIO_2026.xlsx"
ATTRIBUTION  = "lastsign"

CAMPANAS_META = {
    "LA CURACAO": {
        "120234590485200367": "CURACAO_social-media_facebook_conversiones-HOGAR",
        "120234592898620367": "Sebas / CURACAO_social-media_facebook_conversiones-VIDEO",
        "120234601486280367": "CURACAO_social-media_facebook_conversiones-LINEA-BLANCA",
        "120234589433020367": "CURACAO_social-media_facebook_conversiones-TELEFONIA",
        "120235019554320367": "CURACAO_social-media_facebook_conversiones-COMPUTO",
        "120241139755810367": "CURACAO_PISO_facebook_2025_alcance_SOLO_POR_HOY",
        "120240083028940367": "CURACAO_PISO_facebook_alcance_REPOWER_VIDEO",
        "120239176840340367": "CURACAO_PISO_facebook_2025_alcance_REPOWER_LB",
        "120234727712310367": "CURACAO_social-media_facebook_conversiones-PADS",
        "120235716510800367": "CURACAO_PISO_facebook_2025_alcance_REPOWER_TLF",
        "120240447583300367": "CURACAO_PISO_facebook_2025_alcance_REPOWER_COMPUTO",
        "120234715177880367": "CURACAO_social-media_facebook_conversiones-ACCESORIOS",
        "120234606215930367": "CURACAO_social-media_facebook_conversiones-AUDIO",
        "120245549456450367": "CURACAO_PISO_facebook_2026_alcance_CierraPuertas",
        "120240584847660367": "CURACAO_PISO_facebook_2025_alcance_REPOWER_AUDIO",
        "120242107885040367": "CURACAO_PISO_facebook_alcance_REPOWER_PADS",
        "120234669678800367": "CURACAO_social-media_facebook_conversiones-DIGITAL",
        "120251450643280367": "CURACAO_facebook_interaccion_GENERICO",
        "120241194687780367": "CURACAO_social-media_facebook_FEEDOM",
        "120229489114750367": "CURACAO_social-media_facebook_CONECTADOS_conversaciones",
        "120249514628650367": "CURACAO_OMNI_facebook_alcance_CARRUSELES",
        "120245930290960367": "CURACAO_facebook_interaccion_CYBER_sorteo",
        "120240823792620367": "CURACAO_social-media_facebook_INTERACCION_HORATECH",
        "120245809931760367": "CURACAO_social-media_facebook_ALCANCE_HORATECH",
        "120241554555440367": "CURACAO_facebook_conversiones-CATALOGO-HOGAR",
        "120247690081510367": "CURACAO_social-media_facebook_consideration_CYBER",
        "120242155783700367": "CURACAO_PISO_facebook_alcance_BOMBAS_VIDEO",
        "120245610643240367": "CURACAO_OMNI_facebook_interaccion_CierraPuertas_sorteo",
        "120248788311010367": "CURACAO_social-media_facebook_ALCANCE_GENERICO",
        "120244539178720367": "CURACAO_PISO_facebook_2025_alcance_REPOWER_HOGAR",
        "120251459842970367": "CURACAO_social-media_facebook_ALCANCE_ADV-ZTE",
        "120242498510510367": "CURACAO_social-media_facebook_ADV_alcance_MABE_LB",
        "120237406640550367": "CURACAO_social-media_facebook_ADV_alcance_SAMSUNG",
        "120244538917510367": "CURACAO_PISO_facebook_2025_alcance_REPOWER_ACCESORIOS",
        "120236001477860367": "CURACAO_social-media_facebook_ADV_alcance_SAMSUNG_TLF",
        "120241553532050367": "CURACAO_facebook_conversiones-CATALOGO-LINEA-BLANCA",
        "120241554032840367": "CURACAO_facebook_conversiones-CATALOGO-VIDEO",
        "120249501257030367": "CURACAO_social-media_facebook_ADV_alcance_XIAOMI_TLF",
        "120229269369930367": "CURACAO_social-media_facebook_ADV-SAMSUNG_alcance_TELEFONIA",
        "120241609090800367": "CURACAO_PISO_facebook_alcance_REPOWER_MULTICAT",
        "120253945095040367": "CURACAO_social-media_facebook_trafico_ELECTROLUX",
        "120243845421170367": "CURACAO_PISO_facebook_trafico_REPOWER_VIDEO",
        "120245941846580367": "CURACAO_social-media_facebook_ADV_alcance_NVIDIA_COMPUTO",
        "120250555160070367": "CURACAO_facebook_ADV_alcance_AMD_COMPUTO",
        "120243223967760367": "CURACAO_social-media_facebook_ADV_alcance_INTEL_COMPUTO",
        "120241684713590367": "CURACAO_facebook_conversiones-CATALOGO_DIGITAL",
        "120236058386720367": "CURACAO_social-media_facebook_ADV_alcance_SAMSUNG_LB",
        "120237395430930367": "CURACAO_social-media_facebook_ADV_alcance_SAMSUNG_TLF-16.10",
        "120243846791980367": "CURACAO_PISO_facebook_trafico_REPOWER_TLF",
        "120248069439090367": "CURACAO_social-media_facebook_ADV_alcance_HYUNDAI_VIDEO",
        "120243462137640367": "CURACAO_social-media_facebook_ADV_alcance_HISENSE_VIDEO",
        "120246861435760367": "CURACAO_social-media_facebook_ADV_alcance_JBL_AUDIO",
        "120248070369490367": "CURACAO_facebook_conversion_CATALOGO_CYBER",
        "120245980996160367": "CURACAO_social-media_facebook_ADV_alcance_FORLI_HOGAR",
        "120248068609720367": "CURACAO_social-media_facebook_ADV_alcance_INDURAMA_LB",
        "120245980283370367": "CURACAO_social-media_facebook_ADV_alcance_AMD_COMPUTO",
        "120249612862600367": "CURACAO_social-media_facebook_ADV-INDURAMA_alcance_LINEA-BLANCA",
        "120255274922860367": "CURACAO_social-media_facebook_conversiones-remarketing",
        "120248163873260367": "CURACAO_social-media_facebook_ADV_alcance_MRGRILL-HOGAR",
        "120248164842220367": "CURACAO_social-media_facebook_ADV_alcance_CISNE-HOGAR",
        "120251268572330367": "CURACAO_social-media_facebook_ADV_alcance_MABE_LB",
        "120233754325890367": "CURACAO_social-media_facebook_ADV_alcance_GOPRO",
        "120248952315010367": "CURACAO_social-media_facebook_ADV-HYUNDAI_alcance_LINEA-BLANCA",
        "120239949571290367": "CURACAO_social-media_facebook_ADV_alcance_BLACK_FRIDAY",
    },
    "TIENDAS EFE": {
        "120230728497840225": "EFE_social-media_facebook_conversiones-LINEA_BLANCA",
        "120231164329030225": "EFE_social-media_facebook_conversiones-HOGAR",
        "120232753604350225": "EFE_PISO_facebook_2025_alcance_REPOWER_VIDEO",
        "120230729288370225": "EFE_social-media_facebook_conversiones-COMPUTO",
        "120236368406740225": "EFE_PISO_facebook_2025_alcance_SOLO_POR_HOY",
        "120230729117380225": "EFE_social-media_facebook_conversiones-TELEFONIA",
        "120235019214480225": "EFE_PISO_facebook_2025_alcance_REPOWER_LINEA_BLANCA",
        "120236223665690225": "EFE_PISO_facebook_2025_alcance_REPOWER_COMPUTO",
        "120243299525760225": "EFE_facebook_conversiones_catalogo_VIDEO",
        "120230730448300225": "EFE_social-media_facebook_conversiones-VIDEO",
        "120243663201740225": "EFE_social-media_facebook_2026_conversiones-telefonía",
        "120231729539770225": "EFE_PISO_facebook_2025_alcance_REPOWER_TLF",
        "120229645840320225": "EFE_social-media_facebook_2025_conversiones-PADS",
        "120243672117310225": "EFE_social-media_facebook_2026_conversiones-video",
        "120240186851630225": "EFE_PISO_facebook_alcance_CIERRAPUERTAS",
        "120243663913340225": "EFE_social-media_facebook_2026_conversiones-lb",
        "120214902727080225": "COLDEX_social-media_facebook_2025_TRAFICO",
        "120232398213020225": "EFE_social-media_facebook_2025_conversiones-AUDIO",
        "120237661385030225": "EFE_PISO_facebook_alcance_BOMBAS_VIDEO",
        "120213987067410225": "COLDEX_social-media_facebook_2025_01_INTERACCION-LANZAMIENTO",
        "120237627055480225": "EFE_PISO_facebook_alcance_REPOWER_PADS",
        "120235803787860225": "EFE_PISO_facebook_2025_alcance_REPOWER_AUDIO_BW",
        "120226374015830225": "EFE_social-media_facebook_CONECTADOS_conversaciones",
        "120237232793970225": "EFE_facebook_conversiones_CATALOGO_HOGAR",
        "120231560180980225": "EFE_social-media_facebook_conversiones-ACCESORIOS",
        "120237553939830225": "EFE_facebook_conversiones_CATALOGO_PADS",
        "120240512722280225": "EFE_facebook_interaccion__SORTEO_CYBERDAYS",
        "120239384590700225": "EFE_PISO_facebook_2025_alcance_REPOWER_HOGAR",
        "120240219235420225": "EFE_OMNI_facebook_interaccion_CIERRAPUERTAS",
        "120241816731520225": "EFE_social-media_facebook_CONSIDERATION_CYBER",
        "120237248449310225": "EFE_facebook_conversiones_CATALOGO_LINEA-BLANCA",
        "120239384463670225": "EFE_PISO_facebook_2025_alcance_REPOWER_ACCESORIOS",
        "120225163924040225": "EFE_social-media_facebook_ADV-SAMSUNG_alcance_VIDEO",
        "120237184375990225": "EFE_PISO_facebook_2025_alcance_REPOWER_MULTICATEGORIA",
        "120245854008960225": "EFE_social-media_facebook_trafico-ELECTROLUX",
        "120237248515650225": "EFE_facebook_conversiones_CATALOGO_VIDEO",
        "120231729721980225": "EFE_PISO_facebook_2025_trafico_REPOWER_TLF",
        "120238787203240225": "EFE_PISO_facebook_trafico_REPOWER_VIDEO",
        "120225265142040225": "EFE_social-media_facebook_ADV-SAMSUNG_alcance_LINEABLANCA",
        "120231992811170225": "EFE_social-media_facebook_ADV-SAMSUNG_alcance_LB",
        "120243669038550225": "EFE_social-media_facebook_2026_conversiones-apple",
        "120237553648260225": "EFE_facebook_conversiones_CATALOGO_AUDIO",
        "120234508379610225": "EFE_social-media_facebook_ADV-SAMSUNG_alcance_VIDEO - Copia",
        "120242078945160225": "EFE_facebook_conversiones_CATALOGO_RMKT",
        "120227738272320225": "EFE_social-media_facebook_2025_07_conversiones-DIGITAL",
        "120244731141860225": "Nueva campaña de Tráfico - Copia",
        "120244732296660225": "Nueva campaña de Tráfico - Copia (2)",
        "120235751331610225": "EFE_social-media_facebook_ADV-alcance_BLACK_WEEK",
    },
}

MARCAS = {
    "LA CURACAO":  {"contador": "98373248"},
    "TIENDAS EFE": {"contador": "98373144"},
    "JUNTOZ": {"contador": "98373308"}
}



def dia_pertenece_a_evento(d):
    """True/nombre si la fecha cae dentro de CYBER DAYS o CYBER WOW, sin
    importar cuál esté 'activo' hoy. Clave para no perder el último día
    de un evento cuando ya cruzamos al siguiente."""
    for nombre, (ini, fin) in EVENTOS:
        if ini <= d <= fin:
            return nombre
    return None

DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
DIAS_CORTOS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

# --- Ventanas de los cybers de julio 2026 (editables) ---
CYBER_DAYS = (date(2026, 7, 6),  date(2026, 7, 12))   # 6 al 12 julio
CYBER_WOW  = (date(2026, 7, 13), date(2026, 7, 19))   # 13 al 19 julio
EVENTOS = [("CYBER DAYS", CYBER_DAYS), ("CYBER WOW", CYBER_WOW)]


# ---------------------------------------------------------------------------
# MODO — se decide SOLO por la fecha de hoy
# ---------------------------------------------------------------------------
def _hoy_es_cyber(d=None):
    d = d or date.today()
    for _, (ini, fin) in EVENTOS:
        if ini <= d <= fin:
            return True
    return False


ES_CYBER  = _hoy_es_cyber()
ES_DIARIO = not ES_CYBER
MODO_JOB  = os.environ.get("MODO", "auto").strip().lower().rstrip(".")

# En diario JUNTOZ no se consulta ni se publica (es marketplace, no entra en el
# seguimiento diario). El codigo queda intacto: solo cambia que marcas corren.
MARCAS_ACTIVAS = {m: c for m, c in MARCAS.items()
                  if ES_CYBER or m != "JUNTOZ"}


def job_le_toca():
    """True si este workflow debe trabajar hoy. Evita que el job diario y el
    horario se pisen: cada uno se salta los dias del otro."""
    if MODO_JOB == "diario":
        return not ES_CYBER
    if MODO_JOB == "cyber":
        return ES_CYBER
    return True          # 'auto' / sin variable -> corre siempre

# Columna base por marca en el Excel de proyecciones:
#   LA CURACAO : Lunes=B(2) ... Domingo=H(8)
#   TIENDAS EFE: Lunes=K(11) ... Domingo=Q(17)
COL_BASE = {"LA CURACAO": 2, "TIENDAS EFE": 11, "JUNTOZ": 20}


# ---------------------------------------------------------------------------
# PROYECCIONES (Excel, solo CYBER DAYS / CYBER WOW)
# ---------------------------------------------------------------------------
def hoja_evento(hoy=None):
    hoy = hoy or date.today()
    if CYBER_DAYS[0] <= hoy <= CYBER_DAYS[1]:
        return "CYBER DAYS"
    if CYBER_WOW[0] <= hoy <= CYBER_WOW[1]:
        return "CYBER WOW"
    return None


def leer_proyecciones(marca, hoy=None):
    """24 proyecciones (hora 0..23) para la marca y el día dado. Fuera de cyber: ceros."""
    hoy = hoy or date.today()
    hoja = hoja_evento(hoy)
    if hoja is None:
        return [0.0] * 24
    col = COL_BASE[marca] + hoy.weekday()
    wb = load_workbook(ARCHIVO_PROY, data_only=True)
    ws = wb[hoja]
    return [float(ws.cell(row=3 + h, column=col).value or 0) for h in range(24)]

def proy_horaria(marca, fecha=None):
    """24 valores de proyeccion por hora. Solo existe en cyber.
    En diario devuelve ceros y el JSON publica null (no hay curva horaria)."""
    if ES_DIARIO:
        return [0.0] * 24
    return leer_proyecciones(marca, fecha or date.today())


def serie_proy_json(proy):
    """La proyeccion tal como sale al JSON: numeros en cyber, null en diario.
    Publicar null (y no 0) evita que el grafico pinte una linea plana en cero
    y que la tabla muestre una columna de ceros."""
    if ES_DIARIO:
        return [None] * 24
    return [round(v) for v in proy]


def meta_diaria(marca, fecha=None):
    """Meta de sesiones del dia (numero unico).
    diario -> Sheet mensual col F; cyber -> suma de la proyeccion horaria."""
    fecha = fecha or date.today()
    if ES_DIARIO:
        return leer_proyeccion_diaria(marca, fecha)
    return sum(leer_proyecciones(marca, fecha))


# ---------------------------------------------------------------------------
# EXTRACCIÓN YANDEX
# ---------------------------------------------------------------------------
def traer_sesiones_hora(contador):
    """Sesiones/hora de HOY. Devuelve (serie24, corte)."""
    p = {"ids": contador, "metrics": "ym:s:visits",
         "date1": "today", "date2": "today", "group": "hour"}
    r = requests.get(URL_BYTIME, headers=HEADERS, params=p)
    d = r.json()
    if not d.get("data"):
        raise RuntimeError(
            f"Yandex sin 'data' (contador {contador}, HTTP {r.status_code}): "
            f"{d.get('message') or d.get('errors') or d}"
        )
    serie = [int(v or 0) for v in d["data"][0]["metrics"][0]]
    corte = d.get("last_period_index", 23)
    return serie, corte


def traer_sesiones_hora_fecha(contador, fecha):
    """Sesiones/hora de una fecha específica (YYYY-MM-DD). Devuelve serie24."""
    p = {"ids": contador, "metrics": "ym:s:visits",
         "date1": fecha, "date2": fecha, "group": "hour"}
    d = requests.get(URL_BYTIME, headers=HEADERS, params=p).json()
    if not d.get("data"):
        return [0] * 24
    return [int(v or 0) for v in d["data"][0]["metrics"][0]]


def traer_funnel(contador, fecha="today"):
    metrics = ["ym:s:visits", "ym:s:ecommerceRevenue", "ym:s:ecommercePurchases",
               "ym:s:productImpressions", "ym:s:productBasketsQuantity",
               "ym:s:productPurchasedQuantity"]
    p = {"ids": contador, "metrics": ",".join(metrics),
         "date1": fecha, "date2": fecha}
    t = requests.get(URL_DATA, headers=HEADERS, params=p).json()["totals"]
    sesiones, ingresos, compras, vistos, carrito, comprados = t
    return {
        "ingresos": ingresos, "compras": int(compras), "vistos": int(vistos),
        "carrito": int(carrito), "comprados": int(comprados),
        "cr": (compras / sesiones) if sesiones else 0,
        "ticket": (ingresos / compras) if compras else 0,
    }

def evaluar_campanas_numericas(campanas, marca):
    """Recorre la lista de campañas ya traída de Yandex y, si el nombre es
    puro número (Meta/Google sin UTM), lo reemplaza por el nombre real desde
    CAMPANAS_META. Si no está catalogado, lo deja tageado como 'sin catalogar'
    en vez de mostrar el número pelado. JUNTOZ no tiene mapa: queda igual."""
    mapa_marca = CAMPANAS_META.get(marca, {})
    for c in campanas:
        nombre = c["nombre"]
        if RE_CAMPANA_NUMERICA.match(nombre):
            real = mapa_marca.get(nombre)
            c["nombre"] = real if real else f"Facebook Ads (sin catalogar: {nombre})"
    return campanas

def traer_campanas(contador, top=100, fecha="today"):
    p = {"ids": contador, "dimensions": "ym:s:UTMCampaign",
         "metrics": "ym:s:visits,ym:s:ecommercePurchases,ym:s:ecommerceRevenue",
         "date1": fecha, "date2": fecha, "sort": "-ym:s:ecommerceRevenue",
         "limit": top, "attribution": ATTRIBUTION}
    d = requests.get(URL_DATA, headers=HEADERS, params=p).json()
    out = []
    for fila in d.get("data", []):
        nombre = fila["dimensions"][0].get("name") or "(sin campaña)"
        ses, compras, ingresos = fila["metrics"]
        out.append({"nombre": nombre, "sesiones": int(ses),
                    "compras": int(compras), "ingresos": round(ingresos, 2)})
    return out


def traer_top_productos(contador, top=100, fecha="today"):
    p = {"ids": contador, "dimensions": "ym:s:productName,ym:s:productID",
         "metrics": ("ym:s:productImpressions,ym:s:productBasketsQuantity,"
                     "ym:s:productPurchasedQuantity,ym:s:visits,"
                     "ym:s:productPurchasedPrice"),
         "date1": fecha, "date2": fecha,
         "sort": "-ym:s:productPurchasedQuantity", "limit": top}
    d = requests.get(URL_DATA, headers=HEADERS, params=p).json()
    out = []
    for fila in d.get("data", []):
        nombre = fila["dimensions"][0].get("name") or "(sin nombre)"
        sku = fila["dimensions"][1].get("name") or ""
        vistos, carrito, comprados, ses, ingresos = fila["metrics"]
        out.append({"nombre": nombre, "sku": sku, "vistos": int(vistos),
                    "carrito": int(carrito), "comprados": int(comprados),
                    "sesiones": int(ses), "ingresos": round(ingresos, 2)})
    return out


def traer_ingresos_hora(contador, fecha="today"):
    p = {"ids": contador, "metrics": "ym:s:ecommerceRevenue",
         "date1": fecha, "date2": fecha, "group": "hour"}
    d = requests.get(URL_BYTIME, headers=HEADERS, params=p).json()
    if not d.get("data"):
        return [0.0] * 24
    serie = d["data"][0]["metrics"][0]
    return [round(float(v or 0), 2) for v in serie]


# ---------------------------------------------------------------------------
# REPORTE (crudo)
# ---------------------------------------------------------------------------
def armar_reporte(marca):
    cfg = MARCAS[marca]
    dia_idx = date.today().weekday()
    proy = proy_horaria(marca)
    real, corte = traer_sesiones_hora(cfg["contador"])
    funnel = traer_funnel(cfg["contador"])
    real_acum = sum(real[:corte + 1])
    meta = meta_diaria(marca)
    return {
        "marca": marca, "dia": DIAS[dia_idx], "fecha": str(date.today()),
        "corte": corte, "proy": proy, "real": real,
        "meta": meta, "real_acum": real_acum,
        "avance": (real_acum / meta) if meta else 0,
        "funnel": funnel,
    }


# ---------------------------------------------------------------------------
# PAYLOAD data.json
# ---------------------------------------------------------------------------
def construir_marca(marca):
    r = armar_reporte(marca)
    contador = MARCAS[marca]["contador"]
    campanas  = traer_campanas(contador)
    campanas  = evaluar_campanas_numericas(campanas, marca)
    productos = traer_top_productos(contador)
    ing_hora  = traer_ingresos_hora(contador)
    corte = r["corte"]
    proy_json = serie_proy_json(r["proy"])
    por_hora = [{
        "h": h,
        "ses_proy": proy_json[h],
        "ses_real": r["real"][h] if h <= corte else None,
        "ing_real": ing_hora[h]   if h <= corte else None,
    } for h in range(24)]
    f = r["funnel"]
    # meta al corte: solo tiene sentido con curva horaria (cyber).
    # En diario va null y el dashboard compara contra la meta del dia.
    meta_corte = None if ES_DIARIO else round(sum(r["proy"][:corte + 1]))
    bloque = {
        "meta_sesiones_dia":   round(r["meta"]),
        "meta_sesiones_corte": meta_corte,
        "real": {
            "sesiones":  r["real_acum"], "ingresos": round(f["ingresos"], 2),
            "cr": round(f["cr"], 4), "ticket": round(f["ticket"], 2),
            "compras": f["compras"], "vistos": f["vistos"],
            "carrito": f["carrito"], "comprados": f["comprados"],
        },
        "por_hora": por_hora, "campanas": campanas, "top_productos": productos,
    }
    return r, bloque


def construir_payload():
    payload = {
        "generado": datetime.now().astimezone().isoformat(timespec="seconds"),
        "corte": None, "attribution": ATTRIBUTION,
        "modo": "cyber" if ES_CYBER else "diario",
        "evento": hoja_evento(),          # 'CYBER DAYS' / 'CYBER WOW' / None
        "en_vivo": ES_CYBER,              # el dashboard decide el badge con esto
        "hist_retencion_dias": HIST_RETENCION_DIAS,
        "marcas": {}, "semana": {}, "semana_horas": {}, "mes": {},
    }
    reportes = {}
    for marca in MARCAS_ACTIVAS:
        r, bloque = construir_marca(marca)
        payload["marcas"][marca] = bloque
        reportes[marca] = r
        payload["corte"] = r["corte"]
    return payload, reportes


# ---------------------------------------------------------------------------
# SEMANA HORAS — detalle proy/real por hora, por día del cyber, por marca
# ---------------------------------------------------------------------------
def dias_del_evento(hoy=None):
    """(nombre_hoja, [fechas desde el inicio del evento hasta hoy]) o (None, [])."""
    hoy = hoy or date.today()
    hoja = hoja_evento(hoy)
    if hoja is None:
        return None, []
    inicio = CYBER_DAYS[0] if hoja == "CYBER DAYS" else CYBER_WOW[0]
    dias, d = [], inicio
    while d <= hoy:
        dias.append(d)
        d += timedelta(days=1)
    return hoja, dias


def construir_semana_horas(reportes, hoy=None):
    """Para cada marca y cada día del evento: 24 filas {h, proy, real}.
    Hoy usa el real ya extraído (limitado al corte); días pasados se re-consultan
    a Yandex (día completo). Devuelve (nombre_hoja, dict)."""
    hoy = hoy or date.today()
    hoja, dias = dias_del_evento(hoy)
    out = {m: {} for m in MARCAS_ACTIVAS}
    if hoja is None:
        return None, out
    for marca, cfg in MARCAS_ACTIVAS.items():
        cont = cfg["contador"]
        for d in dias:
            proy = proy_horaria(marca, d)
            if d == hoy:
                serie, corte = reportes[marca]["real"], reportes[marca]["corte"]
            else:
                serie, corte = traer_sesiones_hora_fecha(cont, str(d)), 23
            filas = [{
                "h": h,
                "proy": round(proy[h]),
                "real": (serie[h] if h <= corte else None),
            } for h in range(24)]
            out[marca][str(d)] = {"dia": DIAS_CORTOS[d.weekday()], "horas": filas}
    return hoja, out


# ---------------------------------------------------------------------------
# HISTÓRICO (historico.json) — días YA CERRADOS, mismo shape que marcas[x]
# ---------------------------------------------------------------------------
def construir_bloque_fecha(marca, fecha):
    """Bloque de un día CERRADO (corte 23, 24h reales) con el MISMO shape que
    construir_marca(). `fecha` es un objeto date."""
    cont = MARCAS[marca]["contador"]
    fstr = str(fecha)
    proy = proy_horaria(marca, fecha)
    real      = traer_sesiones_hora_fecha(cont, fstr)     # 24h completas
    funnel    = traer_funnel(cont, fstr)
    campanas  = traer_campanas(cont, fecha=fstr)
    campanas  = evaluar_campanas_numericas(campanas, marca)
    productos = traer_top_productos(cont, fecha=fstr)
    ing_hora  = traer_ingresos_hora(cont, fecha=fstr)
    proy_json = serie_proy_json(proy)
    por_hora = [{
        "h": h,
        "ses_proy": proy_json[h],
        "ses_real": real[h],
        "ing_real": ing_hora[h],
    } for h in range(24)]
    meta = round(meta_diaria(marca, fecha))
    return {
        "dia": DIAS_CORTOS[fecha.weekday()],
        "fecha": fstr,
        "meta_sesiones_dia":   meta,
        "meta_sesiones_corte": meta,          # dia cerrado: corte = dia completo
        "real": {
            "sesiones": sum(real), "ingresos": round(funnel["ingresos"], 2),
            "cr": round(funnel["cr"], 4), "ticket": round(funnel["ticket"], 2),
            "compras": funnel["compras"], "vistos": funnel["vistos"],
            "carrito": funnel["carrito"], "comprados": funnel["comprados"],
        },
        "por_hora": por_hora, "campanas": campanas, "top_productos": productos,
    }

def semana_desde_historico(hist, dias=14):
    """Comparativa 'Semana' del diario con días CERRADOS (números finales)."""
    out = {}
    for marca in MARCAS_ACTIVAS:
        dd = hist.get("marcas", {}).get(marca, {})
        filas = []
        for f in sorted(dd.keys())[-dias:]:
            b = dd[f]
            proy = b.get("meta_sesiones_dia", 0)
            real = b.get("real", {}).get("sesiones", 0)
            filas.append({"fecha": f, "dia": b.get("dia", ""),
                          "proy": int(proy), "real": int(real),
                          "avance": (real / proy) if proy else 0})
        out[marca] = filas
    return out


def construir_neto_diario_desde_sheet(hoy=None):
    """Actualiza el bloque 'por_dia' de neto.json con los dias CERRADOS del
    Sheet mensual (col R Venta Real / col O % CR Real / col AL Ticket Real).

    CONTRATO DE ESCRITURA DE neto.json — dos procesos lo tocan:
      · pipeline.py (nube, todos los dias) -> escribe SOLO 'por_dia'.
      · neto.py     (tu PC, solo en cyber) -> escribe SOLO 'hoy' desde el CSV.
    Por eso aqui se hace merge sobre el archivo existente y NUNCA se toca el
    bloque 'hoy': si lo reescribieramos, borrariamos el neto en vivo del cyber.

    Los valores del Sheet se publican TAL CUAL: no se divide entre IGV ni se
    recalcula nada. Lo que ves en la col R es lo que sale en el dashboard.
    """
    hoy = hoy or date.today()
    ayer = hoy - timedelta(days=1)
    inicio = max(HIST_INICIO, hoy - timedelta(days=HIST_RETENCION_DIAS or 60))

    # 1. Punto de partida: lo que ya esta publicado (para no perder 'hoy').
    try:
        with open(ARCHIVO_NETO, "r", encoding="utf-8") as fh:
            prev = json.load(fh)
    except (FileNotFoundError, json.JSONDecodeError):
        prev = {}
    marcas_prev = prev.get("marcas", {}) if isinstance(prev.get("marcas"), dict) else {}

    out = {
        "generado": datetime.now().astimezone().isoformat(timespec="seconds"),
        "fecha_hoy": str(hoy),
        "modo": "cyber" if ES_CYBER else "diario",
        "fuente_dias_cerrados": "sheet_reporte_diario",
        "marcas": {},
    }
    # metadatos de neto.py (corte del neto en vivo) se conservan si existen
    for k in ("corte", "corte_neto", "igv"):
        if k in prev:
            out[k] = prev[k]

    for marca in ("TIENDAS EFE", "LA CURACAO"):
        antes = marcas_prev.get(marca, {}) if isinstance(marcas_prev.get(marca), dict) else {}
        por_dia = {}
        d = inicio
        while d <= ayer:
            n = leer_neto_diario(marca, d)
            if n:
                por_dia[str(d)] = {
                    "venta_neta":  n["venta_neta"],
                    "cr_neto":     n["cr_neto"],
                    "ticket_neto": n["ticket_neto"],
                }
            d += timedelta(days=1)
        bloque = {"por_dia": por_dia}
        if antes.get("hoy") is not None:
            bloque["hoy"] = antes["hoy"]      # intacto: lo maneja neto.py
        out["marcas"][marca] = bloque

    with open(ARCHIVO_NETO, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    return out


# ---------------------------------------------------------------------------
# VISTA MES — acumulado del mes vs meta del mes (suma de la col F del Sheet)
# ---------------------------------------------------------------------------
def construir_mes(hist, reportes, hoy=None):
    """Por marca: meta del mes, real acumulado (dias cerrados del historico +
    lo que va de hoy), proyeccion de cierre y el detalle dia por dia."""
    hoy = hoy or date.today()
    ini_mes = hoy.replace(day=1)
    fin_mes = (ini_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    out = {}
    for marca in MARCAS_ACTIVAS:
        dd = hist.get("marcas", {}).get(marca, {})
        dias, real_acum = [], 0
        d = ini_mes
        while d <= fin_mes:
            key = str(d)
            if d == hoy:
                real = reportes[marca]["real_acum"] if marca in reportes else None
                parcial = True
            elif key in dd:
                real = dd[key].get("real", {}).get("sesiones", 0)
                parcial = False
            else:
                real, parcial = None, False
            if real is not None:
                real_acum += real
            dias.append({"fecha": key, "dia": DIAS_CORTOS[d.weekday()],
                         "proy": round(leer_proyeccion_diaria(marca, d)),
                         "real": real, "parcial": parcial})
            d += timedelta(days=1)

        meta = round(leer_meta_mensual(marca, hoy))
        # Cierre estimado: se cuentan los dias ya cerrados; hoy va aparte porque
        # esta a medio dia y arrastraria el promedio hacia abajo.
        cerrados = [x for x in dias if x["real"] is not None and not x["parcial"]]
        prom = (sum(x["real"] for x in cerrados) / len(cerrados)) if cerrados else 0
        restantes = (fin_mes - hoy).days          # dias despues de hoy
        proy_cierre = round(real_acum + prom * restantes) if prom else None

        out[marca] = {
            "mes": hoy.strftime("%Y-%m"),
            "meta": meta,
            "real": real_acum,
            "avance": (real_acum / meta) if meta else 0,
            "dias_cerrados": len(cerrados),
            "dias_mes": fin_mes.day,
            "proy_cierre": proy_cierre,
            "dias": dias,
        }
    return out


def refrescar_metas_historico(hist):
    """Vuelve a leer del Sheet la meta de sesiones de TODOS los dias guardados.

    Por que: la meta de un dia cerrado se congelaba con el valor que tuviera el
    Sheet la manana siguiente. Si en ese momento el IMPORTRANGE estaba cargando
    o daba #REF!, el dia quedaba con meta 0 para siempre (caso 2026-08-01).
    Releerla en cada corrida es barato -- la pestania del mes ya esta cacheada,
    no cuesta llamadas extra a Yandex -- y ademas recoge cualquier correccion
    que hagas despues en el Sheet.

    Solo pisa el valor guardado cuando el Sheet devuelve algo > 0, para que un
    error transitorio nunca borre una meta buena."""
    cambios = 0
    for marca in MARCAS_ACTIVAS:
        for key, bloque in hist.get("marcas", {}).get(marca, {}).items():
            try:
                nueva = round(leer_proyeccion_diaria(marca, date.fromisoformat(key)))
            except Exception:
                continue
            if nueva > 0 and nueva != bloque.get("meta_sesiones_dia"):
                bloque["meta_sesiones_dia"]   = nueva
                bloque["meta_sesiones_corte"] = nueva
                cambios += 1
    return cambios


def actualizar_historico(hoy=None):
    """Mantiene historico.json con los dias ya CERRADOS.

    · Ventana: desde HIST_INICIO (o los ultimos HIST_RETENCION_DIAS) hasta ayer.
    · 'Ayer' se refresca siempre (Yandex a veces asienta tarde).
    · Los dias ya guardados no se vuelven a consultar.
    · Los faltantes se rellenan de a BACKFILL_MAX por corrida, del mas reciente
      al mas antiguo. Esto evita reventar la cuota de Yandex cuando hay un hueco
      grande (caso tipico: JUNTOZ, que no se consulta en diario, al entrar cyber
      tendria 60 dias vacios de golpe).
    · Fuera de la ventana de retencion se podan los dias viejos.
    """
    hoy = hoy or date.today()
    ayer = hoy - timedelta(days=1)

    inicio = HIST_INICIO
    if HIST_RETENCION_DIAS:
        inicio = max(inicio, hoy - timedelta(days=HIST_RETENCION_DIAS))

    ventana = []
    d = inicio
    while d <= ayer:
        ventana.append(d)
        d += timedelta(days=1)

    hist = {"marcas": {m: {} for m in MARCAS}}
    try:
        with open(SALIDA_HISTORICO, "r", encoding="utf-8") as f:
            prev = json.load(f)
        if isinstance(prev.get("marcas"), dict):
            for m in MARCAS:
                hist["marcas"][m] = prev["marcas"].get(m, {})
    except (FileNotFoundError, json.JSONDecodeError):
        pass

    validos = {str(d) for d in ventana}
    for marca in MARCAS_ACTIVAS:
        guardados = hist["marcas"].setdefault(marca, {})

        # 1. Poda de dias fuera de la ventana de retencion
        for key in [k for k in guardados if k not in validos]:
            del guardados[key]

        # 2. Refresco de los ultimos dias + relleno de faltantes.
        # En cyber corre cada hora, asi que basta con refrescar 'ayer';
        # en diario refrescamos HIST_REFRESH_DIAS porque solo hay una pasada.
        n_ref = HIST_REFRESH_DIAS if ES_DIARIO else 1
        recientes = {str(ayer - timedelta(days=i)) for i in range(n_ref)}

        refrescar  = [d for d in ventana if str(d) in recientes]
        faltantes  = [d for d in reversed(ventana)
                      if str(d) not in guardados and str(d) not in recientes]
        for d in refrescar + faltantes[:BACKFILL_MAX]:
            guardados[str(d)] = construir_bloque_fecha(marca, d)

        faltan = len(faltantes) - BACKFILL_MAX
        if faltan > 0:
            print(f"  {marca}: quedan {faltan} dias por rellenar "
                  f"(se completan en las proximas corridas).")

    # Metas re-leidas del Sheet (arregla dias congelados con meta 0)
    try:
        n = refrescar_metas_historico(hist)
        if n:
            print(f"  metas del historico corregidas desde el Sheet: {n}")
    except Exception as e:
        print(f"  Aviso: no pude refrescar metas del historico ({e}).")

    hist["generado"] = datetime.now().astimezone().isoformat(timespec="seconds")
    with open(SALIDA_HISTORICO, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)
    return hist


# ---------------------------------------------------------------------------
# GOOGLE SHEET "META HORARIA" — escribe la columna Real por día/hora
# ---------------------------------------------------------------------------
# Sheet: "Meta horaria Cyber Julio26"
META_SHEET_KEY = "1kucg2oRhJGWrNET_OdS-AVO9Kdc5NE7jeFJO01mT6oE"
META_TAB = {"CYBER DAYS": "Cyber days", "CYBER WOW": "Cyber wow"}
# Fila de la hora 0 de cada bloque de marca:
META_BASE_ROW = {"LA CURACAO": 3, "TIENDAS EFE": 33, "JUNTOZ": 63}


def escribir_meta_horaria(nombre_evento, semana_horas):
    """Escribe la columna Real de cada día (por marca) en el Sheet Meta horaria.
    Real col = 3 + weekday*4 (C,G,K,O,S,W,AA). VAR/%VAR son fórmulas del Sheet."""
    import gspread
    if not nombre_evento:
        return
    gc = gspread.service_account(filename="credenciales.json")
    ws = gc.open_by_key(META_SHEET_KEY).worksheet(META_TAB[nombre_evento])
    reqs = []
    for marca, dias in semana_horas.items():
        base = META_BASE_ROW[marca]
        for fecha, info in dias.items():
            w = date.fromisoformat(fecha).weekday()
            col = get_column_letter(3 + w * 4)
            valores = [[(fila["real"] if fila["real"] is not None else "")]
                       for fila in info["horas"]]
            reqs.append({"range": f"{col}{base}:{col}{base + 23}", "values": valores})
    if reqs:
        ws.batch_update(reqs, value_input_option="USER_ENTERED")


# ---------------------------------------------------------------------------
# GOOGLE SHEET "BACKEND" (gspread) — pestañas Cortes + Semana (totales/día)
# ---------------------------------------------------------------------------
NOMBRE_SHEET = "CYBER 2026 - BACKEND"


def _abrir_sheet():
    import gspread
    gc = gspread.service_account(filename="credenciales.json")
    return gc.open(NOMBRE_SHEET)


def escribir_corte(reportes):
    ss = _abrir_sheet()
    sh = ss.worksheet("Cortes")
    filas = sh.get_all_values()
    hoy = str(date.today())
    for marca, r in reportes.items():
        hora = f"{r['corte']:02d}:00"
        f = r["funnel"]
        nueva = [marca, hoy, hora, r["real_acum"], round(f["ingresos"], 2),
                 round(f["cr"], 4), round(f["ticket"], 2), round(r["avance"], 4)]
        idx = None
        for i, row in enumerate(filas[1:], start=2):
            if len(row) >= 3 and row[0] == marca and row[1] == hoy and row[2] == hora:
                idx = i
                break
        if idx:
            sh.update(f"A{idx}:H{idx}", [nueva])
        else:
            sh.append_row(nueva, value_input_option="USER_ENTERED")


def escribir_semana(reportes):
    ss = _abrir_sheet()
    sh = ss.worksheet("Semana")
    filas = sh.get_all_values()
    hoy = str(date.today())
    dia = DIAS_CORTOS[date.today().weekday()]
    for marca, r in reportes.items():
        proy_corte = round(sum(r["proy"][:r["corte"] + 1]))
        if ES_DIARIO:
            proy_corte = round(r["meta"])   # sin curva horaria: avance vs meta del día
        nueva = [marca, hoy, dia, round(r["meta"]), r["real_acum"],
                 round(r["real_acum"] / proy_corte, 4) if proy_corte else 0]
        idx = None
        for i, row in enumerate(filas[1:], start=2):
            if len(row) >= 2 and row[0] == marca and row[1] == hoy:
                idx = i
                break
        if idx:
            sh.update(f"A{idx}:F{idx}", [nueva])
        else:
            sh.append_row(nueva, value_input_option="USER_ENTERED")


def leer_semana():
    ss = _abrir_sheet()
    sh = ss.worksheet("Semana")
    filas = sh.get_all_values()[1:]
    out = {}

    def num(x):
        x = str(x).strip().replace(",", ".")
        return float(x) if x else 0.0

    for row in filas:
        if len(row) < 6:
            continue
        marca, fecha, dia, proy, real, avance = row[:6]
        out.setdefault(marca, []).append({
            "fecha": fecha, "dia": dia,
            "proy": int(num(proy)), "real": int(num(real)), "avance": num(avance),
        })
    return out


# ---------------------------------------------------------------------------
# ALERTA DE FALLO
# ---------------------------------------------------------------------------
def alerta(msg):
    try:
        from email.message import EmailMessage
        import smtplib, ssl
        from correo import GMAIL_USER, GMAIL_PASS, DESTINATARIOS
        m = EmailMessage()
        m["Subject"] = f"FALLO pipeline ({'cyber' if ES_CYBER else 'diario'}) - revisar"
        m["From"] = GMAIL_USER
        m["To"] = ", ".join(DESTINATARIOS)
        m.set_content(f"El pipeline ({'cyber' if ES_CYBER else 'diario'}) fallo:\n\n" + msg)
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx) as s:
            s.login(GMAIL_USER, GMAIL_PASS)
            s.send_message(m)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    if not job_le_toca():
        modo_hoy = "cyber" if ES_CYBER else "diario"
        print(f"Hoy es dia {modo_hoy} y este job es MODO={MODO_JOB}. "
              f"No hago nada (lo cubre el otro workflow).")
        return

    print(f"MODO {'CYBER' if ES_CYBER else 'DIARIO'} · "
          f"marcas: {', '.join(MARCAS_ACTIVAS)}")
    print("Extrayendo de Yandex...")
    payload, reportes = construir_payload()

    # --- Sheets BACKEND / META HORARIA: solo tienen sentido en cyber ---
    if ES_CYBER:
        try:
            escribir_corte(reportes)
            escribir_semana(reportes)
            payload["semana"] = leer_semana()
            print("Sheet BACKEND actualizado.")
        except Exception as e:
            print(f"Aviso: Sheet BACKEND no actualizado ({e}). El JSON sigue.")

        try:
            nombre_ev, semana_horas = construir_semana_horas(reportes)
            payload["semana_horas"] = semana_horas
        except Exception as e:
            print(f"Aviso: semana_horas no construida ({e}).")
            nombre_ev, semana_horas = None, {}

        try:
            if nombre_ev:
                escribir_meta_horaria(nombre_ev, semana_horas)
                print("Sheet META HORARIA actualizado.")
        except Exception as e:
            print(f"Aviso: Meta horaria no escrita ({e}). El JSON sigue.")

    # data.json (primera escritura: el dashboard ya tiene lo del dia)
    with open(SALIDA_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"{SALIDA_JSON} escrito.")

    # historico.json — aditivo, NO bloquea el flujo en vivo
    hist = None
    try:
        hist = actualizar_historico()
        print(f"{SALIDA_HISTORICO} actualizado.")
    except Exception as e:
        print(f"Aviso: historico no actualizado ({e}). El resto sigue.")

    # Semana + Mes desde el historico (dias cerrados, numeros finales)
    if hist is not None:
        try:
            payload["semana"] = semana_desde_historico(hist)
            payload["mes"] = construir_mes(hist, reportes)
            with open(SALIDA_JSON, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            print("Semana + Mes reconstruidos desde el historico.")
        except Exception as e:
            print(f"Aviso: semana/mes no reconstruidos ({e}).")

    # neto.json — bloque 'por_dia' desde el Sheet. Nunca toca 'hoy'.
    try:
        construir_neto_diario_desde_sheet()
        print("neto.json actualizado (por_dia desde el Sheet).")
    except Exception as e:
        print(f"Aviso: neto.json no actualizado ({e}).")

    # PNG + correo: solo en cyber
    if ES_CYBER:
        from correo import construir_html, render_png, construir_correo, enviar
        reps_lista = []
        for marca, r in reportes.items():
            render_png(construir_html(r), f"reporte_{marca.replace(' ', '_')}.png")
            reps_lista.append(r)
        enviar(construir_correo(reps_lista))
        print("PNG + correo enviados.")
    else:
        print("Modo diario: sin correo.")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        err = traceback.format_exc()
        print(err, file=sys.stderr)
        alerta(err)          # tambien en diario: si se cae, hay que enterarse
        sys.exit(1)
